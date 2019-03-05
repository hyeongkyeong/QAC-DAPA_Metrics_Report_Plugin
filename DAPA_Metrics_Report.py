#!/usr/bin/env python

"""
    For generating a DAPA Metrics Data Report
    
    Author: Hyeonykyeony Seo

"""

import sys, os, re, datetime, time, codecs, locale
sys.path.insert(0,os.path.join(os.path.dirname(os.path.realpath(__file__)), "lib"))
from optparse import OptionParser
from xml.dom import minidom
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Color, PatternFill, Font, Alignment, Border
import copy
import Metrics_Data_Report


def make_dapa_metrics_report(options):
    input_filename=os.path.dirname(options.output_file)+os.path.sep+"metrics_data.xml"
    metrics_type=['STCYC','STMIF','STPAR','STM29','STCAL','STXLN']
    dapa_metrics_list=parse_dapa_metrics(input_filename,metrics_type)
    gen_excel_file(dapa_metrics_list,metrics_type,options)

def parse_dapa_metrics(xml_file,metrics_type):
    metrics_list=[]
    metrics_tag=ET.parse(xml_file)
    for file_tag in metrics_tag.findall("./File/Entity[@type='function']/.."):
        file_name=file_tag.get('name').split('/')[-1]
        for entity_tag in file_tag.findall("./Entity[@type='function']"):
            function_name=entity_tag.get('name')
            metrics_data_one_function={}
            metrics_data_one_function['file']=file_name
            metrics_data_one_function['function']=function_name
            for metrics_name in metrics_type:
                if entity_tag.find("./Metric[@name='"+metrics_name+"']") != None:
                    metrics_data_one_function[metrics_name]=int(entity_tag.find("./Metric[@name='"+metrics_name+"']").get('value'))
                else:
                    metrics_data_one_function[metrics_name]='None'
            metrics_list.append(metrics_data_one_function)
    return metrics_list
        
def gen_excel_file(metrics_list, metrics_type, options):
    wb=Workbook()
    wsheet=wb.active
    wsheet.title='DAPA_Metrics'
    
    redFill = PatternFill(patternType='solid',fgColor=Color('FFBABA'))
    headerAlignment = Alignment(horizontal='center',vertical='center')
    headerFont = Font(size=12, bold=True)
    datadellAlignment = Alignment(horizontal='center',vertical='center')
    
    wsheet['A1'] = 'File'
    wsheet['B1'] = 'Function'
    wsheet['C1'] = metrics_type[0]
    wsheet['D1'] = metrics_type[1]
    wsheet['E1'] = metrics_type[2]
    wsheet['F1'] = metrics_type[3]
    wsheet['G1'] = metrics_type[4]
    wsheet['H1'] = metrics_type[5]
    wsheet.freeze_panes='A2'
    row=2
    for list_data in metrics_list:
        wsheet['A'+str(row)]=list_data['file']
        wsheet['B'+str(row)]=list_data['function']

        wsheet['C'+str(row)]=list_data['STCYC']
        if list_data['STCYC'] != 'None' and list_data['STCYC'] > 20:
            wsheet['C'+str(row)].fill=redFill
        wsheet['D'+str(row)]=list_data['STMIF']
        if list_data['STMIF'] != 'None' and list_data['STMIF'] > 6:
            wsheet['D'+str(row)].fill=redFill
        wsheet['E'+str(row)]=list_data['STPAR']
        if list_data['STPAR'] != 'None' and list_data['STPAR'] > 8:
            wsheet['E'+str(row)].fill=redFill
        wsheet['F'+str(row)]=list_data['STM29']
        if list_data['STM29'] != 'None' and list_data['STM29'] > 8:
            wsheet['F'+str(row)].fill=redFill
        wsheet['G'+str(row)]=list_data['STCAL']
        if list_data['STCAL'] != 'None' and list_data['STCAL'] > 10:
            wsheet['G'+str(row)].fill=redFill
        wsheet['H'+str(row)]=list_data['STXLN']
        if list_data['STXLN'] != 'None' and list_data['STXLN'] > 200:
            wsheet['H'+str(row)].fill=redFill
        row+=1
        
    for i in range(1,9):
        wsheet.cell(row=1,column=i).alignment=headerAlignment
        wsheet.cell(row=1,column=i).font=headerFont
    for r in range(1,len(metrics_list)+2):
        for c in range(1,9):
            wsheet.cell(row=r,column=c).alignment = datadellAlignment
    wb.save(options.output_file)
    
        
if __name__ == "__main__":
    usage ="""
    Given a resuls data xml file, this script generates a Metrics Data Report.
    Type DAPA_Metrics_Report.py --help for more information.
    """

    # set up proper comand line parsing
    parser = OptionParser(usage=usage, version="%prog 0.1")
    parser.add_option("-d", "--results-data", dest="results_data",
                      help="An xml file containing results data for a QA Framework project.", metavar="FILE")
    parser.add_option("-o", "--output-file", dest="output_file",
                      help="Specify the output file path for the generated document.", metavar="PATH")
    parser.add_option("-i", "--info", action="store_true", dest="info",
                      help="Show report requirements and output format.")

    # parse the command line
    (options, args) = parser.parse_args()

    enc = locale.getpreferredencoding()
    # command line constraints
    if options.info:
        print "require: metrics"
        print "require: cma"
        print "format: xlsx"
        sys.exit(0)
    if not options.results_data:
        parser.error('option --results-data not specified.')
    if not options.output_file:
        parser.error("option --output-file not specified.")
    if not os.path.exists(options.results_data):
        parser.error('the specified results data file could not be found.')

    if enc:
        options.results_data = options.results_data.decode(enc)
        options.output_file = options.output_file.decode(enc)

    # touch the output file to make sure that we can create it.
    try:
        f = open(options.output_file, 'a')
        f.close()
        os.utime(options.output_file, None)
    except:
        print "could not open output file for writing. exiting"
        sys.exit(1)

    options_for_xmlfile=copy.deepcopy(options)
    options_for_xmlfile.output_file=os.path.dirname(options.output_file)+os.path.sep+"metrics_data.xml"
    Metrics_Data_Report.write_xml(options_for_xmlfile)
    make_dapa_metrics_report(options)
